""" This module is used for developing/ accessing data reader utility. """
import os
import logging
import pyexcel as exc
from openpyxl import Workbook
from openpyxl.styles import Font

from Utilities.config_utility import ConfigUtility
import Utilities.logger_utility as log_utils
import openpyxl as xl

from Utilities.report_utility import xls_Report


class DataReader:
    log = log_utils.custom_logger(logging.INFO)
    config = ConfigUtility()

    def __init__(self):
        self.cur_path = os.path.abspath(os.path.dirname(__file__))

    def load_test_data(self):
        records = None
        prop = self.config.load_properties_file()
        base_execution_data = prop.get('RAFT', 'base_execution_data')
        print(base_execution_data)
        ui_file_path = os.path.join(self.cur_path, r"../TestRunner/{}.csv".format(base_execution_data))

        try:
            if ui_file_path is not None:
                records = exc.iget_records(file_name=ui_file_path)
        except Exception as ex:
            raise Exception("Failed to load test data.", ex)

        return records

    def get_data(self, tc_name, column_name):

        value = None
        excel_records = self.load_test_data()

        try:
            if excel_records is not None:
                for record in excel_records:
                    if record['TC_Name'] == tc_name:
                        value = record[column_name]
                        break
                    else:
                        continue

        except Exception as ex:
            raise Exception("Failed to get test data.", ex)

        return value

    def get_cell_data(self, tc_name, column_name):
        value = None
        try:
            ui_file_path = os.getcwd() + '/TestData/TestData_staging.csv'
            if ui_file_path is not None:
                excel_records = exc.iget_records(file_name=ui_file_path)
                if excel_records is not None:
                    for record in excel_records:
                        if record['TC_Name'] == tc_name:

                            value = record[column_name]

                            break
                        else:
                            continue

        except Exception as ex:
            raise Exception("Failed to get test data.", ex)

        return value

    def create_xlsx(self):
        data = [['TC_ID', 'IMMEDIATE RESPONSE', 'STATUS CALLBACK RESPONSE' , 'PASSTHRU RESPONSE', 'CONNECT APPLET '
                                                                                                  'RESPONSE',
                 'DIAL-WHOM RESPONSE',  'GATHER APPLET RESPONSE','EXOTEL CALL DETAILS','STATUS'],]


        wb = Workbook()

        # Accounting format
        fmt_acct = u'_($* #,##0.00_);[Red]_($* (#,##0.00);_($* -_0_0_);_(@'
        for row in data:
            wb.active.append(row)
        font = Font(color="000000")
        ws = wb.active
        ws.freeze_panes = "A2"
        ws["A1"].font = font
        ws["B1"].font = font
        ws.print_title_rows = '1:1'
        file_name = xls_Report()
        wb.save(file_name)
        return file_name

    def get_row_count(self, file):
        try:
            wb = xl.load_workbook(filename=file)
            # the 2 lines under do the same.
            sheet = wb['Sheet']
            # sheet = wb.worksheets[0]

            row_count = sheet.max_row
            column_count = sheet.max_column
        except Exception as e:
            logging.info(e)
            raise Exception
        return row_count, column_count

    def write_data_file(self, file, row, col, text):
        try:
            wb = xl.load_workbook(filename=file)
            # the 2 lines under do the same.
            sheet = wb['Sheet']
            c1 = sheet.cell(row=row + 1, column=col)

            # writing values to cells
            c1.value = text
            wb.save(file)
            wb.close()
        except Exception as e:
            logging.info(e)
            raise Exception

